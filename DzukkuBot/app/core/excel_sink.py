"""
Append confirmed orders / reservations directly to data/Project_Dzukku.xlsx.

Best-effort, file-locked, never raises (failures only logged) so an
Excel-write hiccup never blocks an order from being saved to PostgreSQL.

Targets the existing sheets:
  - "Orders"      : OrderID, customer, Phone, Item, Total, Status, Date/Time,
                    DeliveryDate, Address, Platform, Location_URL, Qty,
                    UnitPrice, InvoiceURL, special
  - "Reservation" : Res_ID, Customer_Name, Phone, Date, Time, Guests, Table_No,
                    Status, Special_Requests, Email
"""

from __future__ import annotations

import logging
import threading
from datetime import datetime
from typing import Iterable

from openpyxl import load_workbook

from app.core.config import settings

logger = logging.getLogger(__name__)

_lock = threading.Lock()


def _open_book():
    return load_workbook(settings.XLSX_PATH)


def _row_for_sheet(ws, headers: list[str], values: dict) -> list:
    """Build a row aligned to the existing header order; missing cols become ''."""
    return [values.get(h, "") for h in headers]


def _headers(ws) -> list[str]:
    return [str(c.value or "").strip() for c in ws[1]]


def append_order(
    order_ref:      str,
    customer_name:  str,
    customer_phone: str,
    items:          Iterable[dict],
    total_price:    float,
    platform:       str = "Telegram",
    status:         str = "Pending",
    special:        str = "",
) -> None:
    if not settings.XLSX_PATH.exists():
        logger.warning("XLSX %s missing — skipping Excel order append.", settings.XLSX_PATH)
        return

    items = list(items)
    pretty = ", ".join(f"{int(i.get('qty', 1))}x {i.get('item_name', i.get('item', '?'))}"
                       for i in items) or "(no items)"
    total_qty = sum(int(i.get("qty", 1)) for i in items)
    unit_price = items[0].get("price", items[0].get("unitPrice", "")) if items else ""

    record = {
        "OrderID":      order_ref,
        "customer":     customer_name,
        "Phone":        customer_phone,
        "Item":         pretty,
        "Total":        total_price,
        "Status":       status,
        "Date/Time":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "DeliveryDate": "",
        "Address":      "",
        "Platform":     platform,
        "Location_URL": "",
        "Qty":          total_qty,
        "UnitPrice":    unit_price,
        "InvoiceURL":   "",
        "special":      special,
    }

    with _lock:
        try:
            wb = _open_book()
            if "Orders" not in wb.sheetnames:
                logger.warning("Sheet 'Orders' not found in %s", settings.XLSX_PATH)
                return
            ws = wb["Orders"]
            headers = _headers(ws)
            ws.append(_row_for_sheet(ws, headers, record))
            wb.save(settings.XLSX_PATH)
            logger.info("Order %s appended to Excel.", order_ref)
        except PermissionError:
            logger.warning(
                "Excel append failed: %s is open in another program. Close it and retry.",
                settings.XLSX_PATH,
            )
        except Exception as e:
            logger.warning("Excel append failed for order %s: %s", order_ref, e)


def append_reservation(
    reservation_ref: str,
    customer_name:   str,
    customer_phone:  str,
    date:            str,
    time:            str,
    guests:          int,
    special_request: str = "",
    table_no:        str = "",
    email:           str = "",
    status:          str = "Confirmed",
) -> None:
    if not settings.XLSX_PATH.exists():
        logger.warning("XLSX %s missing — skipping Excel reservation append.", settings.XLSX_PATH)
        return

    record = {
        "Res_ID":            reservation_ref,
        "Customer_Name":     customer_name,
        "Phone":             customer_phone,
        "Date":              date,
        "Time":              time,
        "Guests":            guests,
        "Table_No":          table_no,
        "Status":            status,
        "Special_Requests":  special_request,
        "Email":             email,
    }

    with _lock:
        try:
            wb = _open_book()
            if "Reservation" not in wb.sheetnames:
                logger.warning("Sheet 'Reservation' not found in %s", settings.XLSX_PATH)
                return
            ws = wb["Reservation"]
            headers = _headers(ws)
            ws.append(_row_for_sheet(ws, headers, record))
            wb.save(settings.XLSX_PATH)
            logger.info("Reservation %s appended to Excel.", reservation_ref)
        except PermissionError:
            logger.warning(
                "Excel append failed: %s is open in another program. Close it and retry.",
                settings.XLSX_PATH,
            )
        except Exception as e:
            logger.warning("Excel append failed for reservation %s: %s",
                           reservation_ref, e)
