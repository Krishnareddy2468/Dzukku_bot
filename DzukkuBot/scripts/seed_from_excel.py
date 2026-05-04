#!/usr/bin/env python3
"""
Seed PostgreSQL from data/Project_Dzukku.xlsx.

Reads every sheet and inserts rows into the corresponding PostgreSQL tables
using the SQLAlchemy ORM models.  Run AFTER alembic upgrade head.

Usage:
    cd DzukkuBot && python scripts/seed_from_excel.py
"""

import asyncio
import hashlib
import logging
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

# ── Path setup so we can import app.* ───────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession  # noqa: E402

from app.core.config import settings  # noqa: E402
from app.db.models import (  # noqa: E402
    Restaurant, User, Customer, Channel, Session as SessionModel,
    MenuCategory, MenuItem, MenuItemImage,
    Cart, CartItem, Order, OrderItem,
    Driver, Delivery, DeliveryLocationEvent,
    Payment, DiningTable, TableSession, TableSessionOrder,
    Reservation, Invoice, OutboxEvent,
)
from app.db.base import Base  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ── Session factory (standalone — not using app.db.session to avoid circulars) ─
engine = create_async_engine(settings.DATABASE_URL, echo=False, pool_size=5)
SessionFactory = async_sessionmaker(bind=engine, class_=AsyncSession, expire_on_commit=False)

# ── Load workbook ────────────────────────────────────────────────────────────
XLSX = ROOT / "data" / "Project_Dzukku.xlsx"
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)


def _rows(sheet_name: str):
    """Yield (header, list-of-dicts) for a sheet."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    data = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            data.append(dict(zip(header, row)))
        else:
            break
    return data


def _hash(plain: str) -> str:
    """bcrypt hash — same as auth routes."""
    import bcrypt
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt()).decode()


# ── Status mapping from Excel → DB ──────────────────────────────────────────

_ORDER_STATUS_MAP = {
    "Delivered": "DELIVERED",
    "Preparing": "PREPARING",
    "Received": "CREATED",
    "Ready": "READY",
    "Out for Delivery": "OUT_FOR_DELIVERY",
    "Cancelled": "CANCELLED",
    "Confirmed": "CONFIRMED",
    "Pending": "CREATED",
}

_RES_STATUS_MAP = {
    "Confirmed": "CONFIRMED",
    "Requested": "CREATED",
    "Cancelled": "CANCELLED",
    "No Show": "NO_SHOW",
}

_EMP_ROLE_MAP = {
    "Restaurant Manager": "MANAGER",
    "Head Chef": "KITCHEN",
    "Sous Chef": "KITCHEN",
    "Line Cook": "KITCHEN",
    "Prep Cook": "KITCHEN",
    "Tandoor Chef": "KITCHEN",
    "Dishwasher": "KITCHEN",
    "Cashier": "CASHIER",
    "Wait Staff": "WAITER",
    "Hostess": "WAITER",
    "Delivery Manager": "MANAGER",
    "Delivery Driver": "DRIVER",
}

_TYPE_MAP = {
    "Veg": "VEG",
    "Non-Veg": "NON_VEG",
}

# ── Category assignments for the 20 primary menu items ──────────────────────

_ITEM_CATEGORIES = {
    "Paneer Butter Masala": "Main Course",
    "Vegetable Biryani": "Rice & Biryani",
    "Aloo Gobi": "Main Course",
    "Chole Bhature": "Breads & Snacks",
    "Palak Paneer": "Main Course",
    "Masala Dosa": "South Indian",
    "Veg Fried Rice": "Rice & Biryani",
    "Veg Manchurian": "Appetizers",
    "Paneer Tikka": "Appetizers",
    "Mushroom Curry": "Main Course",
    "Chicken Biryani": "Rice & Biryani",
    "Mutton Rogan Josh": "Main Course",
    "Butter Chicken": "Main Course",
    "Chicken Tikka": "Appetizers",
    "Fish Curry": "Main Course",
    "Egg Curry": "Main Course",
    "Prawn Masala": "Main Course",
    "Chicken Fried Rice": "Rice & Biryani",
    "Chicken 65": "Appetizers",
    "Grilled Chicken": "Main Course",
}

_CATEGORY_SORT = {
    "Appetizers": 1,
    "South Indian": 2,
    "Breads & Snacks": 3,
    "Main Course": 4,
    "Rice & Biryani": 5,
    "Desserts": 6,
    "Beverages": 7,
    "Combo Meals": 8,
}


# ═══════════════════════════════════════════════════════════════════════════
async def seed():
    async with SessionFactory() as session:
        # ─── 1. Restaurant (create or use existing) ──────────────────────
        result = await session.execute(select(Restaurant).limit(1))
        rest = result.scalar_one_or_none()
        if not rest:
            rest = Restaurant(
                name="Dzukku Restaurant",
                phone="9876543301",
                address="HSR Layout, Bangalore",
                timezone="Asia/Kolkata",
            )
            session.add(rest)
            await session.flush()
        rid = rest.id
        logger.info("Restaurant id=%s", rid)

        # ─── 2. Menu Categories (skip existing) ─────────────────────────
        cat_names = sorted(set(_ITEM_CATEGORIES.values()), key=lambda c: _CATEGORY_SORT.get(c, 99))
        cat_map: dict[str, int] = {}
        # Load existing categories
        existing_cats = (await session.execute(
            select(MenuCategory).where(MenuCategory.restaurant_id == rid)
        )).scalars().all()
        for ec in existing_cats:
            cat_map[ec.name] = ec.id
        for name in cat_names:
            if name in cat_map:
                continue
            mc = MenuCategory(restaurant_id=rid, name=name, sort_order=_CATEGORY_SORT.get(name, 99))
            session.add(mc)
            await session.flush()
            cat_map[name] = mc.id
        logger.info("Categories: %s", cat_map)

        # ─── 3. Menu Items (primary 20 from Master_Menu, skip existing) ─
        menu_rows = _rows("Master_Menu")
        menu_by_name: dict[str, MenuItem] = {}
        # Load existing menu items
        existing_items = (await session.execute(
            select(MenuItem).where(MenuItem.restaurant_id == rid)
        )).scalars().all()
        for ei in existing_items:
            menu_by_name[ei.name.lower()] = ei
        primary_ids = {f"V{i:03d}" for i in range(1, 11)} | {f"NV{i:03d}" for i in range(1, 11)}
        for r in menu_rows:
            item_id_raw = str(r.get("sales dashboard", "") or "").strip()
            if item_id_raw not in primary_ids:
                continue
            name = str(r.get("Item Name", "") or "").strip()
            if name.lower() in menu_by_name:
                continue  # skip existing
            cat_name = _ITEM_CATEGORIES.get(name)
            cat_id = cat_map.get(cat_name) if cat_name else None
            raw_type = str(r.get("Category", "") or "").strip()
            item_type = _TYPE_MAP.get(raw_type, "VEG")
            price = int(float(r.get("Price", 0) or 0)) * 100  # rupees → cents
            sp_raw = r.get("Special_Price")
            sp_cents = int(float(sp_raw) * 100) if sp_raw else None
            available = str(r.get("Status", "") or "").strip() == "Available"
            stock = r.get("Stock")
            stock_int = int(stock) if stock is not None else None
            prep_str = str(r.get("Reciepe", "") or r.get(None, "") or "").strip()
            # Column 10 (index 9) has the prep time string like "15 mins"
            prep_time = 900  # default 15 min
            # Try to parse prep time from the column
            for col_key in list(r.keys()):
                val = r.get(col_key)
                if isinstance(val, str) and "mins" in val.lower():
                    m = re.search(r"(\d+)\s*mins", val, re.I)
                    if m:
                        prep_time = int(m.group(1)) * 60
                    break

            is_special = str(r.get("Is_Special", "") or "").strip().lower() == "yes"
            tags = ["special"] if is_special else None

            mi = MenuItem(
                restaurant_id=rid,
                category_id=cat_id,
                name=name,
                description=str(r.get("Description", "") or "").strip() or None,
                type=item_type,
                price_cents=price,
                special_price_cents=sp_cents,
                available=available,
                stock_qty=stock_int,
                prep_time_sec=prep_time,
                tags=tags,
            )
            session.add(mi)
            await session.flush()
            menu_by_name[name.lower()] = mi
        logger.info("Menu items: %d", len(menu_by_name))

        # ─── 4. Users (Employees, skip existing by email) ────────────────
        emp_rows = _rows("Employees")
        user_map: dict[str, User] = {}  # email → User
        driver_users: list[User] = []
        # Load existing users
        existing_users = (await session.execute(
            select(User).where(User.restaurant_id == rid)
        )).scalars().all()
        for eu in existing_users:
            user_map[eu.email] = eu
            if eu.role == "DRIVER":
                driver_users.append(eu)
        for r in emp_rows:
            name = str(r.get("Name", "") or "").strip()
            role_raw = str(r.get("Role", "") or "").strip()
            role = _EMP_ROLE_MAP.get(role_raw, "KITCHEN")
            phone = str(r.get("Phone", "") or "").strip()
            email = str(r.get("Email", "") or "").strip().lower()
            active = str(r.get("Status", "") or "").strip().lower() == "active"
            if not email:
                continue
            if email in user_map:
                continue  # skip existing
            pw_hash = _hash("dzukku123")  # default password for all seeded staff
            u = User(
                restaurant_id=rid,
                name=name,
                phone=phone,
                email=email,
                password_hash=pw_hash,
                role=role,
                active=active,
            )
            session.add(u)
            await session.flush()
            user_map[email] = u
            if role == "DRIVER":
                driver_users.append(u)
        logger.info("Users: %d (drivers: %d)", len(user_map), len(driver_users))

        # ─── 5. Drivers (skip existing) ─────────────────────────────────
        existing_drivers = (await session.execute(select(Driver))).scalars().all()
        existing_driver_user_ids = {d.user_id for d in existing_drivers}
        for u in driver_users:
            if u.id in existing_driver_user_ids:
                continue
            d = Driver(
                restaurant_id=rid,
                user_id=u.id,
                vehicle_type="BIKE",
                vehicle_no="",
                active=u.active,
            )
            session.add(d)
        await session.flush()
        logger.info("Drivers seeded")

        # ─── 6. Dining Tables (skip existing) ────────────────────────────
        existing_tables = (await session.execute(select(DiningTable))).scalars().all()
        existing_table_names = {t.name for t in existing_tables}
        table_rows = _rows("Tables")
        table_capacity_map = {
            # T01–T04 → 2-seat, T05–T08 → 4-seat, T09–T12 → 6-seat,
            # T13–T16 → 8-seat, T17–T20 → 10-seat
        }
        for tnum in range(1, 21):
            key = f"T{tnum:02d}"
            if key in existing_table_names:
                continue
            if tnum <= 4:
                cap = 2
            elif tnum <= 8:
                cap = 4
            elif tnum <= 12:
                cap = 6
            elif tnum <= 16:
                cap = 8
            else:
                cap = 10
            dt = DiningTable(
                restaurant_id=rid,
                name=key,
                capacity=cap,
                active=True,
            )
            session.add(dt)
        await session.flush()
        logger.info("Tables: 20 seeded")

        # ─── 7. Customers (deduplicated from Orders + Reservations, skip existing) ──
        cust_map: dict[str, Customer] = {}  # phone → Customer
        # Load existing customers
        existing_custs = (await session.execute(
            select(Customer).where(Customer.restaurant_id == rid)
        )).scalars().all()
        for ec in existing_custs:
            cust_map[ec.phone] = ec
        order_rows = _rows("Orders")
        res_rows = _rows("Reservation")
        offline_rows = _rows("Dashboard_Offline")

        # Collect unique customers
        cust_set: dict[str, str] = {}  # phone → name
        for r in order_rows:
            phone = str(r.get("Phone", "") or "").strip()
            name = str(r.get("customer", "") or "").strip()
            if phone and phone not in cust_set:
                cust_set[phone] = name
        for r in res_rows:
            phone = str(r.get("Phone", "") or "").strip()
            name = str(r.get("Customer_Name", "") or "").strip()
            if phone and phone not in cust_set:
                cust_set[phone] = name
        for r in offline_rows:
            phone = str(r.get("Phone", "") or "").strip()
            name = str(r.get("Customer Name", "") or "").strip()
            if phone and phone not in cust_set:
                cust_set[phone] = name

        for phone, name in cust_set.items():
            if phone in cust_map:
                continue
            c = Customer(
                restaurant_id=rid,
                name=name or "Guest",
                phone=phone,
            )
            session.add(c)
            await session.flush()
            cust_map[phone] = c
        logger.info("Customers: %d", len(cust_map))

        # ─── 8. Orders + Order Items ────────────────────────────────────
        # First, collect all item names from orders that aren't in the menu yet
        # and create placeholder menu items for them.
        for r in order_rows:
            item_str = str(r.get("Item", "") or "").strip()
            qty_raw = r.get("Qty")
            unit_price_raw = r.get("UnitPrice")
            if qty_raw and unit_price_raw:
                # Single-item order
                name = item_str.strip()
                if name and name.lower() not in menu_by_name:
                    # Create placeholder menu item
                    try:
                        up = int(float(unit_price_raw) * 100)
                    except (ValueError, TypeError):
                        up = 0
                    mi = MenuItem(
                        restaurant_id=rid,
                        name=name,
                        type="NON_VEG",
                        price_cents=up if up > 0 else 10000,
                        available=True,
                        stock_qty=None,
                        prep_time_sec=900,
                    )
                    session.add(mi)
                    await session.flush()
                    menu_by_name[name.lower()] = mi
            else:
                # Parse "2x Paneer Tikka, 1x Butter Naan" format
                parts = re.split(r",\s*", item_str)
                for part in parts:
                    m = re.match(r"(\d+)x\s+(.+)", part.strip())
                    if m:
                        item_name = m.group(2).strip()
                    else:
                        item_name = part.strip()
                    if not item_name or item_name.lower() in menu_by_name:
                        continue
                    # Create placeholder menu item
                    mi = MenuItem(
                        restaurant_id=rid,
                        name=item_name,
                        type="NON_VEG",
                        price_cents=10000,  # 100 INR default
                        available=True,
                        stock_qty=None,
                        prep_time_sec=900,
                    )
                    session.add(mi)
                    await session.flush()
                    menu_by_name[item_name.lower()] = mi
        await session.flush()
        logger.info("Menu items after adding order-referenced items: %d", len(menu_by_name))

        # Load existing order refs to avoid duplicates
        existing_order_refs = {
            r.order_ref for r in
            (await session.execute(select(Order.order_ref))).scalars().all()
        }
        # Build a mapping of raw Excel OrderID → DB order_ref
        order_count = 0
        seen_order_refs: set[str] = set(existing_order_refs)
        for r in order_rows:
            order_ref_raw = str(r.get("OrderID", "") or "").strip()
            if not order_ref_raw:
                continue
            # Make order_ref unique
            base_ref = f"DZK-{order_ref_raw}"
            order_ref = base_ref
            counter = 1
            while order_ref in seen_order_refs:
                order_ref = f"{base_ref}-{counter}"
                counter += 1
            seen_order_refs.add(order_ref)

            phone = str(r.get("Phone", "") or "").strip()
            customer = cust_map.get(phone)
            customer_id = customer.id if customer else None

            raw_status = str(r.get("Status", "") or "").strip()
            status = _ORDER_STATUS_MAP.get(raw_status, "CREATED")

            total_price = r.get("Total", 0) or 0
            total_cents = int(float(total_price) * 100)

            platform = str(r.get("Platform", "") or "").strip()
            order_type = "DELIVERY"
            if platform in ("Dine-In", "Offline"):
                order_type = "DINE_IN"
            elif platform in ("Cash", "Online"):
                order_type = "DELIVERY"

            dt_val = r.get("Date/Time")
            created_at = dt_val if isinstance(dt_val, datetime) else None

            notes = str(r.get("special ", "") or "").strip() or None

            order = Order(
                restaurant_id=rid,
                order_ref=order_ref,
                customer_id=customer_id,
                order_type=order_type,
                status=status,
                subtotal_cents=total_cents,
                total_cents=total_cents,
                notes=notes,
                created_at=created_at,
            )
            session.add(order)
            await session.flush()

            # Parse item string into individual items
            item_str = str(r.get("Item", "") or "").strip()
            qty_raw = r.get("Qty")
            unit_price_raw = r.get("UnitPrice")

            # If qty/unit_price are provided, create a single OrderItem
            if qty_raw and unit_price_raw:
                try:
                    qty = int(float(qty_raw))
                    unit_cents = int(float(unit_price_raw) * 100)
                except (ValueError, TypeError):
                    qty = 1
                    unit_cents = total_cents
                # Find matching menu item
                mi = menu_by_name.get(item_str.lower())
                item_id_val = mi.id if mi else list(menu_by_name.values())[0].id
                oi = OrderItem(
                    restaurant_id=rid,
                    order_id=order.id,
                    item_id=item_id_val,
                    item_name_snapshot=item_str,
                    qty=qty,
                    unit_price_cents=unit_cents,
                    status="DONE" if status == "DELIVERED" else "PENDING",
                )
                session.add(oi)
            else:
                # Parse "2x Paneer Tikka, 1x Butter Naan" format
                parts = re.split(r",\s*", item_str)
                for part in parts:
                    m = re.match(r"(\d+)x\s+(.+)", part.strip())
                    if m:
                        qty = int(m.group(1))
                        item_name = m.group(2).strip()
                    else:
                        qty = 1
                        item_name = part.strip()
                    if not item_name:
                        continue

                    # Try to match to a menu item
                    mi = menu_by_name.get(item_name.lower())
                    # Fuzzy match
                    if not mi:
                        for key, val in menu_by_name.items():
                            if item_name.lower() in key or key in item_name.lower():
                                mi = val
                                break

                    unit_cents = mi.price_cents if mi else (total_cents // max(qty, 1))
                    item_status = "DONE" if status == "DELIVERED" else "PENDING"

                    oi = OrderItem(
                        restaurant_id=rid,
                        order_id=order.id,
                        item_id=mi.id if mi else list(menu_by_name.values())[0].id,
                        item_name_snapshot=item_name,
                        qty=qty,
                        unit_price_cents=unit_cents,
                        status=item_status,
                    )
                    session.add(oi)
            order_count += 1

        # Also seed Dashboard_Offline orders (deduplicated)
        for r in offline_rows:
            order_ref_raw = str(r.get("OrderID", "") or "").strip()
            if not order_ref_raw:
                continue
            base_ref = f"DZK-OFF-{order_ref_raw}"
            order_ref = base_ref
            counter = 1
            while order_ref in seen_order_refs:
                order_ref = f"{base_ref}-{counter}"
                counter += 1
            seen_order_refs.add(order_ref)

            phone = str(r.get("Phone", "") or "").strip()
            customer = cust_map.get(phone)
            customer_id = customer.id if customer else None

            raw_status = str(r.get("Status", "") or "").strip()
            status = _ORDER_STATUS_MAP.get(raw_status, "CREATED")

            platform = str(r.get("Platform", "") or "").strip()
            order_type = "DINE_IN" if platform in ("Dine-In", "Offline", "Dzukku Restaurant") else "DELIVERY"

            dt_val = r.get("Date/Time")
            created_at = dt_val if isinstance(dt_val, datetime) else None

            order = Order(
                restaurant_id=rid,
                order_ref=order_ref,
                customer_id=customer_id,
                order_type=order_type,
                status=status,
                subtotal_cents=0,
                total_cents=0,
                created_at=created_at,
            )
            session.add(order)
            order_count += 1

        await session.flush()
        logger.info("Orders: %d", order_count)

        # ─── 9. Reservations ────────────────────────────────────────────
        res_count = 0
        seen_res_refs: set[str] = set()
        for r in res_rows:
            res_ref_raw = str(r.get("Res_ID", "") or "").strip()
            if not res_ref_raw:
                continue
            base_ref = f"RSV-{res_ref_raw}"
            res_ref = base_ref
            counter = 1
            while res_ref in seen_res_refs:
                res_ref = f"{base_ref}-{counter}"
                counter += 1
            seen_res_refs.add(res_ref)

            phone = str(r.get("Phone", "") or "").strip()
            customer = cust_map.get(phone)
            customer_id = customer.id if customer else None

            raw_status = str(r.get("Status", "") or "").strip()
            status = _RES_STATUS_MAP.get(raw_status, "CREATED")

            date_val = r.get("Date")
            date_only = date_val.date() if isinstance(date_val, datetime) else None

            time_val = r.get("Time")
            time_only = time_val if hasattr(time_val, "hour") else None

            guests = int(r.get("Guests", 1) or 1)
            special_request = str(r.get("Special_Requests", "") or "").strip() or None

            rv = Reservation(
                restaurant_id=rid,
                reservation_ref=res_ref,
                customer_id=customer_id,
                date=date_only,
                time=time_only,
                guests=guests,
                special_request=special_request,
                status=status,
            )
            session.add(rv)
            res_count += 1
        await session.flush()
        logger.info("Reservations: %d", res_count)

        # ─── 10. Invoices ────────────────────────────────────────────────
        inv_rows = _rows("Invoices")
        inv_headers = list(ws.iter_rows(max_row=1, values_only=True))[0] if (ws := wb["Invoices"]) else []
        # Invoices sheet has no proper header — raw columns
        # Col0: order_ref, Col1: customer_name, Col2: datetime, Col3: URL, Col4: total, Col5: status
        inv_count = 0
        ws_inv = wb["Invoices"]
        for i, row in enumerate(ws_inv.iter_rows(min_row=1, values_only=True)):
            if i == 0 and row[0] and "ORD" in str(row[0]):
                # No header row, first row is data
                pass
            elif i == 0:
                continue
            # Check if it's actually a header (doesn't start with ORD)
            vals = list(row)
            order_ref_raw = str(vals[0] or "").strip() if vals[0] else ""
            if not order_ref_raw or not vals[0]:
                continue

            dt_val = vals[1] if len(vals) > 1 else None
            total_raw = vals[4] if len(vals) > 4 else None
            total_cents = int(float(total_raw) * 100) if total_raw else 0

            created_at_val = vals[2] if len(vals) > 2 else None
            created_at = created_at_val if isinstance(created_at_val, datetime) else None

            inv = Invoice(
                restaurant_id=rid,
                invoice_no=f"INV-{uuid.uuid4().hex[:6].upper()}",
                entity_type="ORDER",
                entity_id=0,
                subtotal_cents=total_cents,
                tax_cents=0,
                total_cents=total_cents,
                pdf_url=str(vals[3] or "").strip() if len(vals) > 3 and vals[3] else None,
                created_at=created_at,
            )
            session.add(inv)
            inv_count += 1
        await session.flush()
        logger.info("Invoices: %d", inv_count)

        # ─── Commit everything ──────────────────────────────────────────
        await session.commit()
        logger.info("=== SEED COMPLETE ===")


if __name__ == "__main__":
    asyncio.run(seed())
